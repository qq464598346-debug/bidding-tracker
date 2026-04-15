"""
Flask REST API服务
为前端网站提供统一的招投标数据接口

API端点:
  GET /api/bidding          - 获取招标列表（支持筛选/搜索/分页）
  GET /api/bidding/:id      - 获取招标详情
  GET /api/stats            - 获取统计数据
  POST /api/crawl           - 手动触发采集
  GET /api/sources          - 获取数据源状态
  GET /api/categories       - 获取分类信息
  GET /api/events           - SSE实时数据推送（WebSocket替代方案）
  GET /api/export/excel     - 导出Excel报告（支持筛选）
  GET /api/export/pdf       - 导出PDF报告（支持筛选）
"""
import asyncio
import json
import time
import threading
import queue
from datetime import datetime
from flask import Flask, jsonify, request, send_from_directory, Response, make_response
from flask_cors import CORS

from core.config import settings
from core.db import Database


# ==================== 全局SSE管理器 ====================
class SSEManager:
    """Server-Sent Events 连接管理器，支持实时推送"""

    def __init__(self):
        self._clients = {}      # {client_id: queue.Queue}
        self._client_counter = 0
        self._lock = threading.Lock()
        # 心跳间隔（秒）
        self._heartbeat_interval = 15
        # 最新数据缓存（用于新连接快速同步）
        self._last_stats = None
        self._last_update_time = None

    def add_client(self):
        """注册新的SSE客户端，返回(client_id, queue)"""
        with self._lock:
            self._client_counter += 1
            client_id = f"client_{self._client_counter}_{int(time.time())}"
            q = queue.Queue()
            self._clients[client_id] = q
            print(f"📡 SSE客户端已连接: {client_id} (当前在线: {len(self._clients)})")
            return client_id, q

    def remove_client(self, client_id):
        """移除断开的SSE客户端"""
        with self._lock:
            if client_id in self._clients:
                del self._clients[client_id]
                print(f"📡 SSE客户端断开: {client_id} (当前在线: {len(self._clients)})")

    def get_client_count(self):
        """获取当前在线客户端数量"""
        with self._lock:
            return len(self._clients)

    def broadcast(self, event_type, data):
        """向所有连接的客户端广播事件"""
        if not self._clients:
            return

        payload = json.dumps({
            'event': event_type,
            'data': data,
            'timestamp': datetime.now().isoformat()
        }, ensure_ascii=False, default=str)

        with self._lock:
            dead_clients = []
            for cid, q in self._clients.items():
                try:
                    q.put_nowait(payload)
                except queue.Full:
                    dead_clients.append(cid)

            for cid in dead_clients:
                del self._clients[cid]

        # 更新缓存
        if event_type in ('data_updated', 'stats_updated'):
            self._last_stats = data
            self._last_update_time = datetime.now().isoformat()

    def broadcast_heartbeat(self):
        """广播心跳包保持连接活跃"""
        self.broadcast('heartbeat', {'time': datetime.now().isoformat()})

    def get_last_stats(self):
        """获取最新统计数据缓存"""
        return self._last_stats, self._last_update_time


# 全局单例
sse_manager = SSEManager()


def create_app():
    """创建Flask应用实例"""
    app = Flask(__name__)
    CORS(app)  # 允许跨域请求

    # ====== 数据库连接管理 ======
    def get_db():
        return Database(settings.DATABASE_PATH)

    # ====== 辅助函数：获取过滤后的数据 ======
    def _get_filtered_data():
        """根据请求参数获取过滤后的招标数据（供导出和SSE使用）"""
        db = get_db()
        try:
            result = db.query(
                keyword=request.args.get('keyword'),
                source=request.args.get('source', 'all'),
                operator=request.args.get('operator', 'all'),
                category=request.args.get('category', 'all'),
                status=request.args.get('status', 'all'),
                min_relevance=float(request.args.get('min_relevance', 0)),
                limit=min(int(request.args.get('limit', 1000)), 5000),
                offset=0,
                sort_by=request.args.get('sort', 'publish_time'),
                order=request.args.get('order', 'desc')
            )
            return result.get('items', [])
        finally:
            db.close()

    # ====== API路由 ======

    @app.route('/')
    def index():
        """API首页 - 返回服务信息"""
        return jsonify({
            "service": "运营商招投标智能监控系统 v4.0",
            "version": "4.0.0",
            "data_sources": {
                "yifangbao": "乙方宝 (yfbzb.com) - 三大运营商招标聚合",
                "xunbiaobao": "百度寻标宝 (xunbiaobao.baidu.com) - 百度招投标平台"
            },
            "features": ["SSE实时推送", "Excel/PDF导出", "定时采集", "乙方宝+寻标宝双数据源"],
            "endpoints": {
                '/api/bidding': 'GET - 招标列表(支持搜索/筛选)',
                '/api/bidding/<int:id>': 'GET - 招标详情',
                '/api/stats': 'GET - 统计数据',
                '/api/crawl': 'POST - 手动触发采集',
                '/api/sources': 'GET - 数据源状态',
                '/api/categories': 'GET - 分类信息',
                '/api/events': 'GET - ⭐ SSE实时数据推送',
                '/api/export/excel': 'GET - ⭐ 导出Excel报告',
                '/api/export/pdf': 'GET - ⭐ 导出PDF报告',
            }
        })

    @app.route('/api/bidding')
    def get_bidding_list():
        """
        获取招标列表
        
        Query参数:
        keyword: 搜索关键词
        source: 数据源 (ggzy/chinamobile/chinaunicom/chinatelecom/all)
        operator: 运营商过滤 (chinamobile/chinaunicom/chinatelecom/all)
        category: 类别 (software/solution/server/service/all)
        status: 状态 (bidding/upcoming/closing_soon/result_ended/all)
        min_relevance: 最低相关度 (0-1, 默认0)
        sort: 排序字段 (publish_time/budget/relevance/deadline)
        order: 排序方向 (asc/desc, 默认desc)
        limit: 每页条数 (默认20, 最大100)
        offset: 偏移量 (默认0)
        """
        db = get_db()

        try:
            result = db.query(
                keyword=request.args.get('keyword'),
                source=request.args.get('source', 'all'),
                operator=request.args.get('operator', 'all'),
                category=request.args.get('category', 'all'),
                status=request.args.get('status', 'all'),
                min_relevance=float(request.args.get('min_relevance', 0)),
                limit=min(int(request.args.get('limit', 20)), 100),
                offset=int(request.args.get('offset', 0)),
                sort_by=request.args.get('sort', 'publish_time'),
                order=request.args.get('order', 'desc')
            )

            return jsonify({
                'code': 200,
                'message': 'success',
                'data': result
            })

        except Exception as e:
            return jsonify({'code': 500, 'message': str(e)}), 500
        finally:
            db.close()

    @app.route('/api/bidding/<int:item_id>')
    def get_bidding_detail(item_id):
        """获取单条招标详情"""
        db = get_db()

        try:
            result = db.query(limit=1, offset=0)

            item = None
            for i in result.get('items', []):
                if i['id'] == item_id:
                    item = i
                    break

            if item:
                return jsonify({'code': 200, 'data': item})
            else:
                return jsonify({'code': 404, 'message': '未找到该记录'}), 404

        except Exception as e:
            return jsonify({'code': 500, 'message': str(e)}), 500
        finally:
            db.close()

    @app.route('/api/stats')
    def get_stats():
        """获取统计数据概览"""
        db = get_db()

        try:
            stats = db.get_stats()

            # 广播统计更新到所有SSE客户端
            sse_manager.broadcast('stats_updated', stats)

            return jsonify({
                'code': 200,
                'data': stats,
                'timestamp': datetime.now().isoformat()
            })
        except Exception as e:
            return jsonify({'code': 500, 'message': str(e)}), 500
        finally:
            db.close()

    @app.route('/api/crawl', methods=['POST'])
    def trigger_crawl():
        """手动触发数据采集（异步执行）+ 广播采集开始事件"""

        def do_crawl():
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                sse_manager.broadcast('crawl_started', {
                    'message': '数据采集中...',
                    'started_at': datetime.now().isoformat()
                })

                from crawlers.coordinator import CrawlCoordinator
                coordinator = CrawlCoordinator(settings)
                results = loop.run_until_complete(coordinator.crawl_all())

                sse_manager.broadcast('crawl_completed', {
                    'message': f'采集完成！共获取 {sum(r.get("count",0) for r in results)} 条数据',
                    'results': results,
                    'completed_at': datetime.now().isoformat()
                })

                # 采集完成后自动广播数据更新
                _broadcast_data_update()
            except Exception as e:
                sse_manager.broadcast('crawl_error', {
                    'message': f'采集出错: {str(e)}',
                    'error_at': datetime.now().isoformat()
                })
            finally:
                loop.close()

        thread = threading.Thread(target=do_crawl, daemon=True)
        thread.start()

        return jsonify({
            'code': 200,
            'message': '采集任务已启动，将通过SSE实时推送进度'
        })

    @app.route('/api/sources')
    def get_sources():
        """获取数据源状态和统计"""
        db = get_db()

        try:
            cursor = db._get_conn().cursor()

            sources_info = []
            for source in ['yifangbao', 'xunbiaobao']:
                cursor.execute("""
                    SELECT source, COUNT(*) as cnt, MAX(crawled_at) as last_crawl
                    FROM bidding_items 
                    WHERE source=?
                    GROUP BY source
                """, (source,))

                row = cursor.fetchone()

                names = {
                    'yifangbao': '乙方宝 (yfbzb.com)',
                    'xunbiaobao': '百度寻标宝 (xunbiaobao.baidu.com)',
                }

                sources_info.append({
                    'id': source,
                    'name': names.get(source, source),
                    'count': row['cnt'] if row else 0,
                    'last_crawl': row['last_crawl'] if row else None
                })

            return jsonify({
                'code': 200,
                'data': sources_info
            })

        finally:
            db.close()

    @app.route('/api/categories')
    def get_categories():
        """获取项目类别定义"""
        categories = [
            {'id': 'software', 'name': '基础软件', 'icon': '💻',
             'keywords': ['操作系统','中间件','数据库','开发平台']},
            {'id': 'solution', 'name': '行业解决方案', 'icon': '💡',
             'keywords': ['智慧城市','5G应用','数字化转型','云计算']},
            {'id': 'server', 'name': '服务器', 'icon': '🖧',
             'keywords': ['服务器','存储','网络设备','GPU算力']},
            {'id': 'service', 'name': '服务', 'icon': '🛠️',
             'keywords': ['运维','集成','安全服务','软件开发']},
            {'id': 'other', 'name': '其他', 'icon': '📋',
             'keywords': []}
        ]

        return jsonify({
            'code': 200,
            'data': categories
        })

    # ================================================================
    #  ★★★ 新增：SSE 实时数据推送端点 ★★★
    # ================================================================

    @app.route('/api/events')
    def sse_stream():
        """
        Server-Sent Events 实时数据流
        
        支持的事件类型:
          - connected:     连接确认
          - data_updated:  新数据到达（含最新招标条目）
          - stats_updated: 统计数据变化
          - crawl_started:  采集任务启动通知
          - crawl_completed: 采集任务完成通知
          - crawl_error:   采集错误通知
          - heartbeat:     心跳保活（每15秒）
        
        使用方式:
          const es = new EventSource('/api/events');
          es.onmessage = (e) => { ... };
          es.addEventListener('data_updated', (e) => { ... });
        """
        client_id, msg_queue = sse_manager.add_client()

        def generate():
            """SSE生成器函数 - 持续发送事件"""
            # 发送连接确认 + 初始数据快照
            yield f"event: connected\ndata: {json.dumps({'client_id': client_id, 'connected_at': datetime.now().isoformat()}, ensure_ascii=False)}\n\n"

            # 发送最新的统计数据作为初始同步
            last_stats, last_time = sse_manager.get_last_stats()
            if last_stats:
                yield f"event: stats_updated\ndata: {json.dumps(last_stats, ensure_ascii=False, default=str)}\n\n"

            # 持续监听消息队列
            try:
                while True:
                    try:
                        # 阻塞等待最多30秒
                        msg = msg_queue.get(timeout=30)
                        event_type = json.loads(msg).get('event', 'message')
                        yield f"event: {event_type}\ndata: {msg}\n\n"
                    except queue.Empty:
                        # 超时发送注释行保持连接（SSE规范）
                        yield ": heartbeat keep-alive\n\n"
            except GeneratorExit:
                # 客户端断开
                pass
            finally:
                sse_manager.remove_client(client_id)

        response = Response(
            generate(),
            mimetype='text/event-stream',
            headers={
                'Cache-Control': 'no-cache',
                'Connection': 'keep-alive',
                'X-Accel-Buffering': 'no',  # 禁用Nginx缓冲
                'Access-Control-Allow-Origin': '*',
            }
        )
        return response

    @app.route('/api/events/status')
    def sse_status():
        """查看当前SSE连接状态"""
        return jsonify({
            'code': 200,
            'data': {
                'online_clients': sse_manager.get_client_count(),
                'service': 'SSE real-time push',
                'endpoint': '/api/events',
                'supported_events': [
                    'connected', 'data_updated', 'stats_updated',
                    'crawl_started', 'crawl_completed', 'crawl_error', 'heartbeat'
                ]
            }
        })

    # ================================================================
    #  ★★★ 新增：数据导出端点（Excel + PDF）★★★
    # ================================================================

    @app.route('/api/export/excel')
    def export_excel():
        """
        导出招投标数据为 Excel 文件 (.xlsx)
        
        Query参数(与/api/bidding一致):
          keyword, operator, source, category, status, sort, order
        
        返回: Excel文件下载
        """
        items = _get_filtered_data()

        if not items:
            return jsonify({'code': 404, 'message': '没有符合条件的数据可导出'}), 404

        try:
            import io
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '招投标信息'

            # 定义样式
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # 运营商颜色映射
            op_fills = {
                'chinamobile': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
                'chinaunicom': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
                'chinatelecom': PatternFill(start_color='E0F7FA', end_color='E0F7FA', fill_type='solid'),
                'yifangbao': PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid'),
                'xunbiaobao': PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid'),
            }

            cat_map_cn = {'software': '基础软件', 'solution': '行业解决方案', 'server': '服务器', 'service': '服务', 'other': '其他'}
            op_map_cn = {'chinamobile': '中国移动', 'chinaunicom': '中国联通', 'chinatelecom': '中国电信',
                         'yifangbao': '乙方宝', 'xunbiaobao': '寻标宝'}
            status_map_cn = {'bidding': '招标中', 'upcoming': '即将开始', 'closing_soon': '即将截止', 'result_published': '已公示', 'ended': '已结束'}

            headers = ['序号', '数据源', '运营商', '项目名称', '项目编号', '类别', '状态',
                       '预算金额(万元)', '发布时间', '截止时间', '采购方', '区域', 'AI相关度', 'AI标签']

            # 写表头
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 写数据
            for row_idx, item in enumerate(items, 2):
                row_data = [
                    row_idx - 1,
                    op_map_cn.get(item.get('source', ''), item.get('source', '')),
                    op_map_cn.get(item.get('operator', ''), item.get('operator', '')),
                    item.get('title', ''),
                    item.get('project_code', ''),
                    cat_map_cn.get(item.get('category', ''), item.get('category', '')),
                    status_map_cn.get(item.get('status', ''), item.get('status', '')),
                    item.get('budget', 0),
                    item.get('publish_time', ''),
                    item.get('deadline', ''),
                    item.get('purchaser', ''),
                    item.get('region', ''),
                    round(item.get('ai_relevance_score', 0), 2),
                    ', '.join(item.get('ai_tags', []) or [])
                ]

                op = item.get('operator', '')
                row_fill = op_fills.get(op, None)

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=(col_idx in [4, 13]))
                    if row_fill and col_idx > 1:
                        cell.fill = row_fill

            # 设置列宽
            col_widths = [6, 12, 10, 50, 18, 12, 10, 14, 18, 18, 20, 12, 10, 30]
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # 冻结首行
            ws.freeze_panes = 'A2'

            # 自动筛选
            ws.auto_filter.ref = ws.dimensions

            # 写入内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"招投标数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = make_response(output.read())
            response.headers['Content-Type'] = \
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = \
                f'attachment; filename="{filename.encode("utf-8").decode("latin-1")}"'

            print(f"📊 Excel导出完成: {filename} ({len(items)}条记录)")
            return response

        except ImportError:
            return jsonify({
                'code': 500,
                'message': 'Excel导出需要安装openpyxl库: pip install openpyxl'
            }), 500
        except Exception as e:
            return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500

    @app.route('/api/export/pdf')
    def export_pdf():
        """
        导出招投标数据为 PDF 报告
        
        Query参数(与/api/bidding一致):
          keyword, operator, source, category, status, sort, order
        
        返回: PDF文件下载（包含封面、摘要、明细表）
        """
        items = _get_filtered_data()

        if not items:
            return jsonify({'code': 404, 'message': '没有符合条件的数据可导出'}), 404

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm, cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os

            # 尝试注册中文字体
            font_registered = False
            font_paths = [
                "C:/Windows/Fonts/msyh.ttc",    # 微软雅黑 Windows
                "C:/Windows/Fonts/simhei.ttf",   # 黑体 Windows
                "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",  # Linux
                "/System/Library/Fonts/PingFang.ttc",  # macOS
            ]
            font_name = 'ChineseFont'
            for fp in font_paths:
                if os.path.exists(fp):
                    try:
                        pdfmetrics.registerFont(TTFont(font_name, fp))
                        font_registered = True
                        break
                    except Exception:
                        continue

            if not font_registered:
                font_name = 'Helvetica'  # 回退到默认字体

            output_path = settings.DATA_DIR / f"export_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

            doc = SimpleDocTemplate(
                str(output_path),
                pagesize=A4,
                rightMargin=15*mm, leftMargin=15*mm,
                topMargin=20*mm, bottomMargin=15*mm
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle', parent=styles['Title'],
                fontName=font_name, fontSize=22, spaceAfter=6*mm,
                textColor=colors.HexColor('#1a237e')
            )
            subtitle_style = ParagraphStyle(
                'Subtitle', parent=styles['Normal'],
                fontName=font_name, fontSize=11, textColor=colors.grey,
                spaceAfter=10*mm, alignment=1  # 居中
            )
            heading_style = ParagraphStyle(
                'Heading', parent=styles['Heading2'],
                fontName=font_name, fontSize=14, spaceBefore=8*mm, spaceAfter=4*mm,
                textColor=colors.HexColor('#1565c0')
            )
            body_style = ParagraphStyle(
                'Body', parent=styles['Normal'],
                fontName=font_name, fontSize=9, leading=14
            )
            small_style = ParagraphStyle(
                'Small', parent=styles['Normal'],
                fontName=font_name, fontSize=8, leading=11
            )

            story = []

            # ========== 封面 ==========
            story.append(Spacer(1, 30*mm))
            story.append(Paragraph("运营商招投标信息分析报告", title_style))
            story.append(Spacer(1, 5*mm))

            now_str = datetime.now().strftime('%Y年%m月%d日 %H:%M')
            story.append(Paragraph(f"生成时间：{now_str}", subtitle_style))
            story.append(Spacer(1, 10*mm))

            # ========== 数据概要 ==========
            total_budget = sum(it.get('budget', 0) or 0 for it in items)
            cat_counts = {}
            op_counts = {}
            for it in items:
                c = it.get('category', 'other')
                o = it.get('operator', 'unknown')
                cat_counts[c] = cat_counts.get(c, 0) + 1
                op_counts[o] = op_counts.get(o, 0) + 1

            cat_map_cn = {'software': '基础软件', 'solution': '行业解决方案', 'server': '服务器', 'service': '服务', 'other': '其他'}
            op_map_cn = {'chinamobile': '中国移动', 'chinaunicom': '中国联通', 'chinatelecom': '中国电信',
                         'yifangbao': '乙方宝', 'xunbiaobao': '寻标宝'}

            summary_data = [
                ['指标', '数值'],
                ['总记录数', f'{len(items)} 条'],
                ['涉及预算总额', f'{total_budget/10000:.2f} 亿元' if total_budget >= 10000 else f'{total_budget:.0f} 万元'],
            ]
            for k, v in sorted(op_counts.items(), key=lambda x:-x[1]):
                summary_data.append([op_map_cn.get(k, k), f'{v} 条'])
            for k, v in sorted(cat_counts.items(), key=lambda x:-x[1]):
                summary_data.append([cat_map_cn.get(k, k), f'{v} 条'])

            story.append(Paragraph("📊 数据概要", heading_style))
            summary_table = Table(summary_data, colWidths=[80*mm, 60*mm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a73e8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(summary_table)

            # ========== 明细表格 ==========
            story.append(PageBreak())
            story.append(Paragraph("📋 招标明细清单", heading_style))
            story.append(Spacer(1, 3*mm))

            table_headers = ['序号', '运营商', '项目名称', '类别', '状态', '预算(万)', '截止日期']
            table_data = [table_headers]
            for idx, it in enumerate(items[:50], 1):  # 最多50条
                title = it.get('title', '')[:35] + ('...' if len(it.get('title',''))>35 else '')
                table_data.append([
                    str(idx),
                    op_map_cn.get(it.get('operator', ''), ''),
                    title,
                    cat_map_cn.get(it.get('category', ''), ''),
                    it.get('status', ''),
                    f"{it.get('budget', 0):,.0f}" if it.get('budget') else '-',
                    str(it.get('deadline', ''))[:10] if it.get('deadline') else '-',
                ])

            col_w = [15*mm, 20*mm, 70*mm, 22*mm, 20*mm, 22*mm, 25*mm]
            detail_table = Table(table_data, colWidths=col_w, repeatRows=1)
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 7.5),
                ('ALIGN', (0, 0), (1, -1), 'CENTER'),
                ('ALIGN', (4, 4), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#dadce0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fafbff')]),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                # 超长文本换行
                ('WRAP', (2, 1), (2, -1), True),
            ]))
            story.append(detail_table)

            if len(items) > 50:
                story.append(Spacer(1, 5*mm))
                story.append(Paragraph(
                    f"<i>注：本报告仅展示前50条记录，完整共{len(items)}条。请使用Excel导出获取全部数据。</i>",
                    small_style
                ))

            # ========== 页脚说明 ==========
            story.append(Spacer(1, 10*mm))
            story.append(Paragraph(
                "<i>— 本报告由运营商招投标智能监控系统 v4.0 自动生成 —</i><br/>"
                "<i>数据来源：乙方宝(yfbzb.com) + 百度寻标宝(xunbiaobao.baidu.com)</i><br/>"
                f"<i>生成时间：{now_str}</i>",
                ParagraphStyle('FooterNote', parent=small_style, alignment=1, textColor=colors.grey)
            ))

            doc.build(story)

            # 返回文件
            filename = f"招投标报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            with open(output_path, 'rb') as f:
                file_data = f.read()

            response = make_response(file_data)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = \
                f'attachment; filename="{filename.encode("utf-8").decode("latin-1")}"'

            print(f"📄 PDF导出完成: {filename} ({len(items)}条记录)")
            return response

        except ImportError:
            return jsonify({
                'code': 500,
                'message': 'PDF导出需要安装reportlab库: pip install reportlab'
            }), 500
        except Exception as e:
            return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500

    def _broadcast_data_update():
        """广播数据更新事件给所有SSE客户端（内部调用）"""
        db = get_db()
        try:
            latest = db.query(limit=5, offset=0, sort_by='publish_time', order='desc')
            sse_manager.broadcast('data_updated', {
                'new_count': len(latest.get('items', [])),
                'latest_items': [it['title'] for it in latest.get('items', [])],
                'message': f'有 {len(latest.get("items", []))} 条新数据可用'
            })
        except Exception:
            pass
        finally:
            db.close()

    # ====== 静态文件代理（可选，用于前端部署）=====
    @app.route('/<path:path>')
    def serve_static(path):
        """提供前端静态文件"""
        static_dir = settings.ROOT_DIR.parent / 'frontend'  # 前端目录
        if static_dir.exists() and (static_dir / path).exists():
            return send_from_directory(str(static_dir), path)
        return jsonify({'code': 404, 'message': 'Not Found'}), 404

    return app


class APIServer:
    """API服务器包装类"""

    def __init__(self, settings):
        self.settings = settings
        self.app = create_app()
        self._heartbeat_thread = None
        self._running = False

    def _start_heartbeat(self):
        """后台心跳线程，定期发送SSE心跳和检查采集调度"""
        def heartbeat_loop():
            while self._running:
                try:
                    sse_manager.broadcast_heartbeat()
                except Exception:
                    pass
                time.sleep(sse_manager._heartbeat_interval)

        self._heartbeat_thread = threading.Thread(target=heartbeat_loop, daemon=True)
        self._heartbeat_thread.start()

    def run(self):
        """启动服务器"""
        self._running = True
        self._start_heartbeat()

        print(f"\n{'='*56}")
        print(f"  🌐 运营商招投标智能监控系统 v4.0")
        print(f"  数据源: 乙方宝 + 百度寻标宝")
        print(f"  地址: http://{self.settings.API_HOST}:{self.settings.API_PORT}")
        print(f"{'─'*56}")
        print(f"  ✨ 功能:")
        print(f"    📡 SSE实时推送: /api/events")
        print(f"    📊 Excel导出:   /api/export/excel")
        print(f"    📄 PDF报告:     /api/export/pdf")
        print(f"    🕷️ 爬虫采集:    POST /api/crawl (6个数据源并发)")
        print(f"{'='*56}\n")

        self.app.run(
            host=self.settings.API_HOST,
            port=self.settings.API_PORT,
            debug=self.settings.API_DEBUG,
            use_reloader=False,  # 避免重复启动
            threaded=True
        )
